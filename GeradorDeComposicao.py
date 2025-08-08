import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from PIL import Image as PILImage
import io
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import re

# Função para redimensionar uma imagem para caber em um bloco específico da planilha
def resize_image(img_path, max_width=350, max_height=100):
    """Redimensiona a imagem para caber no bloco A1:B5"""
    try:
        pil_img = PILImage.open(img_path)
        ratio = min(max_width/pil_img.width, max_height/pil_img.height)
        new_width = int(pil_img.width * ratio)
        new_height = int(pil_img.height * ratio)
        pil_img = pil_img.resize((new_width, new_height), PILImage.LANCZOS)
        img_byte_arr = io.BytesIO()
        pil_img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)
        return XLImage(img_byte_arr)
    except Exception as e:
        print(f"Erro ao redimensionar imagem: {e}")
        return None

# Função para formatar o dia do vencimento a partir de uma célula da planilha
def formatar_dia_vencimento(vencimento_cell):
    """Retorna o dia do vencimento como string, igual ao valor da planilha de apuração"""
    if pd.isna(vencimento_cell) or vencimento_cell == "":
        return "N/D"
    if hasattr(vencimento_cell, "day"):
        return str(vencimento_cell.day).zfill(2)
    if isinstance(vencimento_cell, (int, float)):
        if 1 <= int(vencimento_cell) <= 31:
            return str(int(vencimento_cell)).zfill(2)
        else:
            return str(int(vencimento_cell))
    vencimento_str = str(vencimento_cell).strip()
    match = re.match(r'^(\d{1,2})[/-]', vencimento_str)
    if match:
        return match.group(1).zfill(2)
    if vencimento_str.isdigit() and 1 <= int(vencimento_str) <= 31:
        return vencimento_str.zfill(2)
    return vencimento_str

# Função para copiar a formatação de uma linha modelo para outra linha
def copiar_formatacao_linha(ws, linha_origem, linha_destino, col_inicio=1, col_fim=13):
    for col in range(col_inicio, col_fim + 1):
        letra = get_column_letter(col)
        cell_origem = ws[f"{letra}{linha_origem}"]
        cell_destino = ws[f"{letra}{linha_destino}"]
        cell_destino.font = cell_origem.font.copy()
        cell_destino.border = cell_origem.border.copy()
        cell_destino.fill = cell_origem.fill.copy()
        cell_destino.number_format = cell_origem.number_format
        cell_destino.protection = cell_origem.protection.copy()
        cell_destino.alignment = cell_origem.alignment.copy()
    ws.row_dimensions[linha_destino].height = ws.row_dimensions[linha_origem].height

# Função principal que gera as composições a partir do arquivo Excel selecionado
def gerar_composicoes(file_path, periodo, imagem_path=None):
    output_path = "composicoes_geradas.xlsx"
    apuracao_df = pd.read_excel(file_path, sheet_name="apuração exemplo")
    apuracao_df.columns = apuracao_df.columns.str.strip()
    apuracao_df['Local Prest.Serviço II'] = apuracao_df['Local Prest.Serviço II'].astype(str).str.strip()
    wb_saida = load_workbook(file_path)
    for wsname in wb_saida.sheetnames:
        if wsname != "ModeloComposicao":
            del wb_saida[wsname]
    gerou = False
    for filial in apuracao_df['Local Prest.Serviço II'].unique():
        filial_str = str(filial).strip()
        dados_filial = apuracao_df[apuracao_df['Local Prest.Serviço II'] == filial_str]
        if dados_filial.empty:
            continue
        gerou = True
        ws_copia = wb_saida.copy_worksheet(wb_saida["ModeloComposicao"])
        ws_copia.title = f"Filial_{filial_str}"
        if imagem_path and os.path.exists(imagem_path):
            try:
                img = resize_image(imagem_path)
                if img:
                    img.anchor = 'A1'
                    ws_copia.add_image(img)
                    ws_copia.merge_cells('A1:B4')
                    for row in range(1, 5):
                        ws_copia.row_dimensions[row].height = 20
                    ws_copia.column_dimensions['A'].width = 25
                    ws_copia.column_dimensions['B'].width = 25
            except Exception as e:
                print(f"Não foi possível adicionar a imagem: {e}")
        ws_copia["B6"] = dados_filial['CNPJ FILIAL'].iloc[0]
        ws_copia["B7"] = filial_str
        ws_copia["B8"] = dados_filial['Local Prest. Serviço'].iloc[0]
        ws_copia["B9"] = "ISS RETIDO"
        ws_copia["B10"] = periodo
        vencimento_cell = dados_filial['Vencimento'].iloc[0]
        dia_vencimento = formatar_dia_vencimento(vencimento_cell)
        ws_copia["B11"] = dia_vencimento
        max_notas = len(dados_filial)
        modelo_linha = 15
        for linha in range(15, 15 + max_notas):
            copiar_formatacao_linha(ws_copia, modelo_linha, linha)
            for col in range(1, 14):
                ws_copia[f"{get_column_letter(col)}{linha}"] = ""
        for i, (_, row) in enumerate(dados_filial.iterrows()):
            linha_excel = 15 + i
            copiar_formatacao_linha(ws_copia, modelo_linha, linha_excel)
            ws_copia[f"A{linha_excel}"] = row['NÚM. DOC. CONTÁBIL']
            ws_copia[f"B{linha_excel}"] = row['Nº NF']
            ws_copia[f"C{linha_excel}"] = row['Base ISS']
            aliq = row['Aliquota']
            try:
                aliq_float = float(str(aliq).replace(",", "."))
                aliq_fmt = f"{int(round(aliq_float * 100))}%" if aliq_float < 1 else f"{int(round(aliq_float))}%"
            except Exception:
                aliq_fmt = str(aliq)
            ws_copia[f"D{linha_excel}"] = aliq_fmt
            ws_copia[f"E{linha_excel}"] = row['Multa']
            ws_copia[f"F{linha_excel}"] = row['Juros']
            ws_copia[f"G{linha_excel}"] = row['Taxa de Emissão']
            ws_copia[f"H{linha_excel}"] = row['ISS Retido']
            ws_copia[f"I{linha_excel}"] = row['CNPJ Prestador']
            ws_copia[f"J{linha_excel}"] = row['Data documento']
            ws_copia[f"K{linha_excel}"] = row['Data de lançamento']
            ws_copia[f"L{linha_excel}"] = row['Município Prestador']
            ws_copia[f"M{linha_excel}"] = row['Cód. Serviços']

        linha_totais = 15 + max_notas + 1
        for i in range(5):
            copiar_formatacao_linha(ws_copia, modelo_linha, linha_totais + i)

        total_iss = pd.to_numeric(dados_filial['ISS Retido'], errors='coerce').fillna(0).sum()
        total_multa = pd.to_numeric(dados_filial['Multa'], errors='coerce').fillna(0).sum()
        total_juros = pd.to_numeric(dados_filial['Juros'], errors='coerce').fillna(0).sum()
        total_taxa = pd.to_numeric(dados_filial['Taxa de Emissão'], errors='coerce').fillna(0).sum()
        total_geral = total_iss + total_multa + total_juros + total_taxa

        ws_copia[f"A{linha_totais}"] = "Total ISS"
        ws_copia[f"B{linha_totais}"] = total_iss
        ws_copia[f"A{linha_totais+1}"] = "Total Multa"
        ws_copia[f"B{linha_totais+1}"] = total_multa
        ws_copia[f"A{linha_totais+2}"] = "Total Juros"
        ws_copia[f"B{linha_totais+2}"] = total_juros
        ws_copia[f"A{linha_totais+3}"] = "Total Taxa"
        ws_copia[f"B{linha_totais+3}"] = total_taxa
        ws_copia[f"A{linha_totais+4}"] = "Total Geral"
        ws_copia[f"B{linha_totais+4}"] = total_geral

        bold_font = Font(bold=True)
        formato_contabil = '_-"R$"* #,##0.00_-;-"R$"* -#,##0.00_-;_-"R$"* "-"??_-;_-@_-'
        for i in range(5):
            ws_copia[f"A{linha_totais+i}"].font = bold_font
            ws_copia[f"B{linha_totais+i}"].number_format = formato_contabil

        # Espaço para informações adicionais: linha azul escuro, depois área de 6 linhas mescladas
        linha_info = linha_totais + 4 + 3
        ws_copia[f"A{linha_info}"] = "Informações adicionais:"
        ws_copia[f"A{linha_info}"].font = Font(bold=True, color="FFFFFF")  # Negrito e branco
        ws_copia[f"A{linha_info}"].fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        ws_copia.merge_cells(f"A{linha_info}:M{linha_info}")

        linha_area = linha_info + 1
        ws_copia.merge_cells(f"A{linha_area}:M{linha_area+5}")

    if "ModeloComposicao" in wb_saida.sheetnames and len(wb_saida.sheetnames) > 1:
        del wb_saida["ModeloComposicao"]
    wb_saida.save(output_path)
    return gerou

def selecionar_arquivo():
    file_path = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx")]
    )
    if not file_path:
        return
    periodos = [f"{mes:02d}.{ano}" for ano in range(2020, 2031) for mes in range(1, 13)]
    periodo_win = tk.Toplevel(root)
    periodo_win.title("Selecione o Período")
    periodo_win.geometry("350x200")
    tk.Label(periodo_win, text="Selecione o período:", font=("Arial", 12)).pack(pady=10)
    periodo_var = tk.StringVar()
    combo = ttk.Combobox(periodo_win, textvariable=periodo_var, values=periodos, state="readonly", font=("Arial", 11))
    combo.pack(pady=5)
    combo.current(0)
    imagem_path = None
    def selecionar_imagem():
        nonlocal imagem_path
        imagem_path = filedialog.askopenfilename(
            title="Selecione a imagem (opcional)",
            filetypes=[("Imagens", "*.png;*.jpg;*.jpeg;*.bmp;*.gif"), ("Todos os arquivos", "*.*")]
        )
        if imagem_path:
            img_label.config(text=f"Imagem selecionada: {imagem_path.split('/')[-1]}")
        else:
            img_label.config(text="Nenhuma imagem selecionada.")
    img_btn = tk.Button(periodo_win, text="Selecionar imagem (opcional)", command=selecionar_imagem, font=("Arial", 10))
    img_btn.pack(pady=5)
    img_label = tk.Label(periodo_win, text="Nenhuma imagem selecionada.", font=("Arial", 9))
    img_label.pack()
    def confirmar():
        periodo = periodo_var.get()
        if not periodo:
            messagebox.showerror("Erro", "Período não selecionado.")
            return
        try:
            gerou = gerar_composicoes(file_path, periodo, imagem_path)
            if gerou:
                messagebox.showinfo("Sucesso", "Composições geradas com sucesso!\nArquivo: composicoes_geradas.xlsx")
            else:
                messagebox.showwarning("Aviso", "Nenhuma composição foi gerada para o arquivo selecionado.")
            periodo_win.destroy()
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro:\n{e}")
    btn_ok = tk.Button(periodo_win, text="Gerar Composições", command=confirmar, font=("Arial", 11), bg="#4CAF50", fg="white")
    btn_ok.pack(pady=10)

root = tk.Tk()
root.title("Gerador de Composições ISS")
root.geometry("470x260")
root.resizable(False, False)
frame = tk.Frame(root, padx=20, pady=20)
frame.pack(expand=True)
label = tk.Label(frame, text="Gerador de Composições ISS", font=("Arial", 16, "bold"))
label.pack(pady=(0, 10))
desc = tk.Label(frame, text="1. Clique no botão abaixo\n2. Selecione o arquivo Excel\n3. Escolha o período desejado\n4. (Opcional) Selecione uma imagem para inserir\n5. O arquivo será gerado na mesma pasta.", font=("Arial", 11))
desc.pack(pady=(0, 15))
btn = tk.Button(frame, text="Selecionar arquivo e gerar", command=selecionar_arquivo, height=2, width=32, font=("Arial", 12), bg="#1976D2", fg="white")
btn.pack()
root.mainloop()